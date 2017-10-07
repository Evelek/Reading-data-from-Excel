import openpyxl, os

class Excel:
    def __init__(self, file_name, spreadsheet_name):
        self.wb = openpyxl.load_workbook(file_name)
        self.sheet = self.wb.get_sheet_by_name(spreadsheet_name)

    def _check_path(self, dirname):
        if os.path.exists(dirname):
            i = 1
            while os.path.exists(dirname + str(i)):
               i += 1
            dirname += str(i)
        os.mkdir(dirname)
        os.chdir(dirname)

    def _print_data(self, i, File):
        print('{0:^5}|{1:^14}|{2:^16}|{3:^13}|{4:^8}|{5:^7}'.format(
            i-1,
            self.sheet.cell(row=i,column=2).value, 
            self.sheet.cell(row=i, column=3).value, 
            self.sheet.cell(row=i, column=4).value, 
            self.sheet.cell(row=i, column=5).value, 
            self.sheet.cell(row=i, column=6).value))
        File.write('{0:^5}|{1:^14}|{2:^16}|{3:^13}|{4:^8}|{5:^7}'.format(
            i-1,
            self.sheet.cell(row=i,column=2).value, 
            self.sheet.cell(row=i, column=3).value, 
            self.sheet.cell(row=i, column=4).value, 
            self.sheet.cell(row=i, column=5).value, 
            self.sheet.cell(row=i, column=6).value))
        File.write('\n')

    def print_all(self):
        self._check_path('Whole_spreadsheet')

        File = open('spreadsheet.txt', 'w')
        print('{0:^5}|{1:^14}|{2:^16}|{3:^13}|{4:^8}|{5:^7}'.format('Lp', 'Imie', 'Nazwisko', 'Miejscowosc', 'Brutto', 'Netto'))
        File.write('{0:^5}|{1:^14}|{2:^16}|{3:^13}|{4:^8}|{5:^7}\n'.format('Lp', 'Imie', 'Nazwisko', 'Miejscowosc', 'Brutto', 'Netto'))
        for i in range(2, self.sheet.max_row + 1):
            self._print_data(i, File)
        File.close()
        print()
        os.chdir('..')
   
    def printAllFromCities(self):
        # collect all cities
        cities = []
        for i in range(2, self.sheet.max_row + 1):
            if self.sheet.cell(row=i, column=4).value not in cities:
                cities.append(self.sheet.cell(row=i, column=4).value)

        self._check_path('By_city')

        for city in cities:
            File = open('{}.txt'.format(city), 'w')
            print('{0:^5}|{1:^14}|{2:^16}|{3:^13}|{4:^8}|{5:^7}'.format('Lp', 'Imie', 'Nazwisko', 'Miejscowosc', 'Brutto', 'Netto'))
            File.write('{0:^5}|{1:^14}|{2:^16}|{3:^13}|{4:^8}|{5:^7}\n'.format('Lp', 'Imie', 'Nazwisko', 'Miejscowosc', 'Brutto', 'Netto'))
            for i in range(2, self.sheet.max_row + 1):
                if(self.sheet.cell(row=i, column=4).value == city):
                    self._print_data(i, File)
        File.close()
        print()
        os.chdir('..')

    def printFromCity(self, city):
        self._check_path('Selected_city')

        File = open('{}.txt'.format(city), 'w')
        print('{0:^5}|{1:^14}|{2:^16}|{3:^13}|{4:^8}|{5:^7}'.format('Lp', 'Imie', 'Nazwisko', 'Miejscowosc', 'Brutto', 'Netto'))
        File.write('{0:^5}|{1:^14}|{2:^16}|{3:^13}|{4:^8}|{5:^7}\n'.format('Lp', 'Imie', 'Nazwisko', 'Miejscowosc', 'Brutto', 'Netto'))
        for i in range(2, self.sheet.max_row + 1):
            if(self.sheet.cell(row=i, column=4).value == city):
                self._print_data(i, File)
        File.close()
        print()
        os.chdir('..')

ex = Excel('Zarobki.xlsx', 'ZAROBKI')
ex.print_all()
ex.printAllFromCities()
selected_city = input('Select city: ')
ex.printFromCity(selected_city)
